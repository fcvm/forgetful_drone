import subprocess
import shutil
import cv2
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from matplotlib.patches import Polygon
from matplotlib.patches import FancyBboxPatch
from matplotlib.patches import BoxStyle

from matplotlib.lines import Line2D
import matplotlib.ticker as mtick

import numpy as np
import config

import os

import pathlib

import multiprocessing
import time

from tqdm import tqdm
import torch

import pandas as pd
import openpyxl

import json


from typing import List, Dict





"""lr_over_epochs = []
def valid_keypoints_plot(batch_image, batch_outputs, batch_labels, epoch_i, ds_name, batch_i, lr):


    batch_size = len(batch_image.detach().cpu())

    for sample_i in range(batch_size):

        # detach the image, keypoints, and output tensors from GPU to CPU
        # just get a single datapoint from each batch
        image = batch_image.detach().cpu()[sample_i]
        image = image.cpu().detach().numpy()
        image = np.transpose(image, (1, 2, 0))
        image = image.copy()
        outputs = batch_outputs.detach().cpu().numpy()[sample_i]
        labels = batch_labels.detach().cpu().numpy()[sample_i]
        
        # Get the specific stuff
        if config.LABELS_USE_WAYPOINT:
            output_wp_x__ = outputs[0]
            output_wp_y__ = outputs[1]
            labels_wp_x__ = labels[0]
            labels_wp_y__ = labels[1]
            if config.LABELS_USE_SPEED:
                output_v = outputs[2]
                labels_v = labels[2]
                if config.LABELS_USE_CTRL_CMD:
                    pass #x[:, 3:] = x[:, 3:]
            else:
                if config.LABELS_USE_CTRL_CMD:
                    pass #x[:, 2:] = x[:, 2:]
        else:
            if config.LABELS_USE_SPEED:
                output_v = outputs[0]
                labels_v = labels[0]
                if config.LABELS_USE_CTRL_CMD:
                    pass #x[:, 1:] = x[:, 1:]
            else:
                if config.LABELS_USE_CTRL_CMD:
                    pass #x[:, 0:] = x[:, 0:]
        #output_wp_x__ = outputs[0]
        #output_wp_y__ = outputs[1]
        #output_v = outputs[2]

        #labels_wp_x__ = labels[0]
        #labels_wp_y__ = labels[1]
        #labels_v = labels[2]

        if config.LABELS_USE_WAYPOINT:

            # and scale them
            output_wp_x = int(config.IMAGE_WIDTH_RESIZED / 2 * (1 + output_wp_x__))
            output_wp_y = int(config.IMAGE_HEIGHT_RESIZED / 2 * (1 + output_wp_y__))
            
            labels_wp_x = int(config.IMAGE_WIDTH_RESIZED / 2 * (1 + labels_wp_x__))
            labels_wp_y = int(config.IMAGE_HEIGHT_RESIZED / 2 * (1 + labels_wp_y__))
        


            # COLORS AND ALPHA
            back_c = "white"
            back_a = 0.6
            label_c = "cyan"
            output_c = "magenta"
            lr_c = "orange"

        
        
        
        
            # plot waypoint distance in pixels
            wp_pxl_dist = ((output_wp_x - labels_wp_x)**2 + (output_wp_y - labels_wp_y)**2)**0.5
            
            
            label_output_distance_plt, = plt.plot(
                [output_wp_x, labels_wp_x],[output_wp_y, labels_wp_y], 
                c=back_c, alpha=back_a, label=f"Pixel dist.: {wp_pxl_dist:.3f}"
            )

            
            # plot labels wp
            label_wp_plt = plt.scatter(labels_wp_x, labels_wp_y, c=label_c, marker="1", s=500)

            # plot prediction wp
            output_wp_plt = plt.scatter(output_wp_x, output_wp_y, c=output_c, marker="2", s=500)

            handles = [
                Line2D([0], [0], ms=12, linestyle='none', mfc=label_c, mec=label_c, marker='1', label=f"Label: x={labels_wp_x__:.2f}, y={labels_wp_y__:.2f}"),
                Line2D([0], [0], ms=12, linestyle='none', mfc=output_c, mec=output_c, marker='2', label=f"Output: x={output_wp_x__:.2f}, y={output_wp_y__:.2f}")]
            l1 = plt.legend(handles=handles, loc=1)

        

        if config.LABELS_USE_SPEED:

            # plot speeds
            speed_bar_x1 = config.IMAGE_WIDTH_RESIZED * 0.01
            speed_bar_x2 = config.IMAGE_WIDTH_RESIZED * 0.09
            speed_bar_y1 = config.IMAGE_HEIGHT_RESIZED * 0.02
            speed_bar_y2 = config.IMAGE_HEIGHT_RESIZED * 0.98

            speed_bar_width = speed_bar_x2 - speed_bar_x1
            speed_bar_height = speed_bar_y2 - speed_bar_y1

            speed_bar_xm = speed_bar_x1 + speed_bar_width/2
            speed_bar_ym = speed_bar_y1 + speed_bar_height/2
            

            speed_bar_plt = plt.gca().add_patch(FancyBboxPatch(
                (speed_bar_x1, speed_bar_y1), speed_bar_width, speed_bar_height,
                boxstyle=BoxStyle("Round", pad=0.3, rounding_size=5*config.IMAGE_RESIZE_FACTOR), mutation_scale=1,
                #mutation_aspect=4,
                fc=back_c, alpha=back_a))


            triang_label_ym = speed_bar_y2 - speed_bar_height * labels_v
            triang_label_ym = min([triang_label_ym, speed_bar_y2])
            triang_label_ym = max([triang_label_ym, speed_bar_y1])

            triang_label_y1 = triang_label_ym - speed_bar_width/4
            triang_label_y2 = triang_label_ym + speed_bar_width/4
            
            pt1 = (speed_bar_xm, triang_label_ym)
            pt2 = (speed_bar_x2, triang_label_y1)
            pt3 = (speed_bar_x2, triang_label_y2)

            label_speed_plt = plt.gca().add_patch(Polygon(
                np.array([pt1, pt2, pt3]),
                facecolor=label_c, label=f"Label, speed {labels_v:.2f}"))



            triang_output_ym = speed_bar_y2 - speed_bar_height * output_v
            triang_output_ym = min([triang_output_ym, speed_bar_y2])
            triang_output_ym = max([triang_output_ym, speed_bar_y1])

            triang_output_y1 = triang_output_ym - speed_bar_width/4
            triang_output_y2 = triang_output_ym + speed_bar_width/4
            
            pt1 = (speed_bar_xm, triang_output_ym)
            pt2 = (speed_bar_x1, triang_output_y1)
            pt3 = (speed_bar_x1, triang_output_y2)

            output_speed_plt = plt.gca().add_patch(Polygon(
                np.array([pt1, pt2, pt3]),
                facecolor=output_c, label=f"Output, speed: {output_v:.2f}"))

            handles = [
                Line2D([0], [0], ms=10, linestyle='none', mfc=label_c, mec=label_c, marker='<', label=f"Label, v: {labels_v:.2f}"),
                Line2D([0], [0], ms=10, linestyle='none', mfc=output_c, mec=output_c, marker='>', label=f"Output, v: {output_v:.2f}")]
            l2 = plt.legend(handles=handles, loc=4)

            if config.LABELS_USE_WAYPOINT:
                plt.gca().add_artist(l1)

        
        plt.imshow(image)
        plt.grid(b=None)
        plt.axis('off')
        #plt.xticks(np.arange(11)/10 * config.DATA_IMAGE_WIDTH_RESIZED, ['-1.0','-0.8','-0.6','-0.4','-0.2','0.0','0.2','0.4','0.6','0.8','1.0'])  # Set text labels.
        #plt.yticks(np.arange(11)/10 * config.DATA_IMAGE_HEIGHT_RESIZED, ['-1.0','-0.8','-0.6','-0.4','-0.2','0.0','0.2','0.4','0.6','0.8','1.0'])
        #plt.title(f"Epoch {epoch_i}, Val. Dataset {ds_name}")
        DIR_PATH = f"{config.OUTPUT_DIR}/validation_visualization/{ds_name}/epoch_{epoch_i:04d}/batch_{batch_i:04d}/"
        FILE_PATH = DIR_PATH + f"sample_{sample_i:04d}.png"

        if not os.path.isdir(DIR_PATH):
            os.makedirs(DIR_PATH)

        plt.savefig(FILE_PATH, bbox_inches='tight', pad_inches=0)
        
        if sample_i == batch_size - 1:

            
            lr_over_epochs.append(lr[0])

            lr_graph_x_max = epoch_i / config.EPOCH_N * config.IMAGE_WIDTH_RESIZED
            lr_graph_x = np.linspace(0, lr_graph_x_max, len(lr_over_epochs))

            lr_graph_y_max = config.IMAGE_HEIGHT_RESIZED
            lr_graph_y_min = config.IMAGE_HEIGHT_RESIZED / 2
            lr_graph_y = [lr_graph_y_max - x / config.LEARN_RATE_INITIAL * (lr_graph_y_max - lr_graph_y_min) for x in lr_over_epochs]

            lr_over_epochs_plt = plt.plot(lr_graph_x, lr_graph_y, c=lr_c)
            plt.text(lr_graph_x[-1], lr_graph_y[-1], f"{lr_over_epochs[-1]:.6f}", c=lr_c)
            
            

            DIR_PATH = f"{config.OUTPUT_DIR}/validation_visualization/{ds_name}/batch_{batch_i:04d}_sample_{sample_i:04d}_over_epochs/"
            FILE_PATH = DIR_PATH + f"epoch_{epoch_i:04d}.png"

            if not os.path.isdir(DIR_PATH):
                os.makedirs(DIR_PATH)
            
            plt.savefig(FILE_PATH, bbox_inches='tight', pad_inches=0)    

        #plt.savefig(f"{config.OUTPUT_DIR}/{ds_name}/{epoch_i}/dataset_{ds_name}_epoch_{epoch_i}_sample_{sample_i}.png", bbox_inches='tight', pad_inches=0)
        plt.close()
"""





def generate_validation_run_video (
    epoch_i: int,
    run_id: str,
    run_dloader: torch.utils.data.DataLoader, 
    outputs_over_samples: list, 
    loss_over_batches: list,
    img_height: int,
    img_width: int,
    frame_rate: float, #config.IMAGE_FRAMES_PER_SECOND
    dpath: pathlib.Path,
) -> None:
    """
    This function ...
      - IMU data hard to visualize -> no impl. so far
      - Ctrl cmd: no impl. so far
    """





    # PARAMETERS
    wp_marker_arm_length = img_height / 20
    wp_rf_axis_length = img_height / 8
    bg_alpha = 0.6
    bg_color = (255,255,255) #"white"
    lbl_color = (255,255,0) #"cyan"
    out_color = (255,0,255) #"magenta"
    loss_color = (0,255,255) #"yellow"
    lbl_out_distance_color = (255, 0, 0) #blue
    fg_line_thickness = 2
    bg_line_thickness = 1
    text_line_thickness = 1

    # CREATE VIDEO WRITER
    [x.unlink() for x in dpath.iterdir() if x.is_file() and f"valid_{run_id}_epoch_" in x.name]
    video_file = dpath/f"valid_{run_id}_epoch_{epoch_i:04d}.mp4"
    video_codec = cv2.VideoWriter_fourcc(*'mp4v')
    video_fps = frame_rate
    video_frame_size = (img_width, img_height)
    video_writer = cv2.VideoWriter(str(video_file), video_codec, video_fps, video_frame_size)


    # CREATE AND WRITE VIDEO FRAMES
    batch_n = len(run_dloader)
    with tqdm(enumerate(run_dloader), leave=False) as pbar:
        pbar.set_description(f"|      Video")
        
        for batch_i, batch in pbar:
            
            batch_img = list(batch['inputs']['image'].numpy())
            
            batch_lbl = list(batch['label'].numpy())
            batch_wp = [x[:2] for x in batch_lbl]
            batch_v = [x[2] for x in batch_lbl]

            batch_size = len(batch_img)

            for batch_sample_i in range(batch_size):
                run_sample_i = batch_size * batch_i + batch_sample_i

                img = batch_img[batch_sample_i]
                if len(img.shape) == 4: 
                    img = img[-1]
                img = np.transpose(img, (1, 2, 0))
                img = img.copy()
                img *= 255
                img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)

                out = outputs_over_samples[run_sample_i]
                
                
                # Fetch values
                lbl_wp = batch_wp[batch_sample_i]
                lbl_wp_x = lbl_wp[0]
                lbl_wp_y = lbl_wp[1]
                out_wp_x = out[0]
                out_wp_y = out[1]
                
                # Scale waypoints for plotting on image
                lbl_wp_x_inPxl = int(img_width / 2 * (1 + lbl_wp_x))
                lbl_wp_y_inPxl = int(img_height / 2 * (1 + lbl_wp_y))
                out_wp_x_inPxl = int(img_width / 2 * (1 + out_wp_x))
                out_wp_y_inPxl = int(img_height / 2 * (1 + out_wp_y))


                # PLOT: Waypoint Reference Frame
                wp_rf_x1 = int(img_width / 2 - wp_rf_axis_length)
                wp_rf_x2 = int(img_width / 2 + wp_rf_axis_length)
                wp_rf_y1 = int(img_height / 2)
                wp_rf_y2 = int(img_height / 2)
                img = cv2.line(img, 
                    (wp_rf_x1, wp_rf_y1), 
                    (wp_rf_x2, wp_rf_y2), 
                    color=bg_color, thickness=bg_line_thickness
                    )

                wp_rf_x1 = int(img_width / 2)
                wp_rf_x2 = int(img_width / 2)
                wp_rf_y1 = int(img_height / 2 - wp_rf_axis_length)
                wp_rf_y2 = int(img_height / 2 + wp_rf_axis_length)
                img = cv2.line(img, 
                    (wp_rf_x1, wp_rf_y1), 
                    (wp_rf_x2, wp_rf_y2), 
                    color=bg_color, thickness=bg_line_thickness
                    )
                
                # PLOT: Waypoint distance
                img = cv2.line(img, 
                    (out_wp_x_inPxl, out_wp_y_inPxl), 
                    (lbl_wp_x_inPxl, lbl_wp_y_inPxl), 
                    color=lbl_out_distance_color, thickness=fg_line_thickness
                )

                # PLOT: Label Waypoint
                img = cv2.line(img, 
                    (lbl_wp_x_inPxl, lbl_wp_y_inPxl), 
                    (lbl_wp_x_inPxl, lbl_wp_y_inPxl + int(wp_marker_arm_length)), 
                    color=lbl_color, thickness=fg_line_thickness
                    )
                img = cv2.line(img, 
                    (lbl_wp_x_inPxl, lbl_wp_y_inPxl), 
                    (
                        lbl_wp_x_inPxl + int(wp_marker_arm_length * np.cos(30 / 180 * np.pi)), 
                        lbl_wp_y_inPxl - int(wp_marker_arm_length * np.sin(30 / 180 * np.pi))
                    ), 
                    color=lbl_color, thickness=fg_line_thickness
                    )
                img = cv2.line(img, 
                    (lbl_wp_x_inPxl, lbl_wp_y_inPxl), 
                    (
                        lbl_wp_x_inPxl - int(wp_marker_arm_length * np.cos(30 / 180 * np.pi)), 
                        lbl_wp_y_inPxl - int(wp_marker_arm_length * np.sin(30 / 180 * np.pi))
                    ), 
                    color=lbl_color, thickness=fg_line_thickness
                    )

                # PLOT: Output Waypoint
                img = cv2.line(img, 
                    (out_wp_x_inPxl, out_wp_y_inPxl), 
                    (out_wp_x_inPxl, out_wp_y_inPxl - int(wp_marker_arm_length)), 
                    color=out_color, thickness=fg_line_thickness
                    )
                img = cv2.line(img, 
                    (out_wp_x_inPxl, out_wp_y_inPxl), 
                    (
                        out_wp_x_inPxl + int(wp_marker_arm_length * np.cos(30 / 180 * np.pi)), 
                        out_wp_y_inPxl + int(wp_marker_arm_length * np.sin(30 / 180 * np.pi))
                    ), 
                    color=out_color, thickness=fg_line_thickness
                    )
                img = cv2.line(img, 
                    (out_wp_x_inPxl, out_wp_y_inPxl), 
                    (
                        out_wp_x_inPxl - int(wp_marker_arm_length * np.cos(30 / 180 * np.pi)), 
                        out_wp_y_inPxl + int(wp_marker_arm_length * np.sin(30 / 180 * np.pi))
                    ), 
                    color=out_color, thickness=fg_line_thickness
                    )
                

            
                # Fetch values
                lbl_speed = batch_v[batch_sample_i]
                out_speed = out[2]

                # PLOT: Speed Bar
                speed_bar_x1 = int(img_height * 0.05)
                speed_bar_x2 = int(img_width * 0.05) + speed_bar_x1
                speed_bar_y1 = int(img_height * 0.05)
                speed_bar_y2 = int(img_height * 0.95)

                speed_bar_w = int(speed_bar_x2 - speed_bar_x1)
                speed_bar_h = int(speed_bar_y2 - speed_bar_y1)

                speed_bar_xc = int(speed_bar_x1 + speed_bar_w/2)
                speed_bar_yc = int(speed_bar_y1 + speed_bar_h/2)

                img = cv2.rectangle(img, 
                    (speed_bar_x1, speed_bar_y1), 
                    (speed_bar_x2, speed_bar_y2), 
                    color=bg_color, thickness=bg_line_thickness
                )

                # PLOT: Speed distance
                lbl_speed_tri_yc = int(speed_bar_y2 - speed_bar_h * lbl_speed)
                lbl_speed_tri_yc = min([lbl_speed_tri_yc, speed_bar_y2])
                lbl_speed_tri_yc = max([lbl_speed_tri_yc, speed_bar_y1])

                out_speed_tri_yc = speed_bar_y2 - int(speed_bar_h * out_speed)
                out_speed_tri_yc = min([out_speed_tri_yc, speed_bar_y2])
                out_speed_tri_yc = max([out_speed_tri_yc, speed_bar_y1])

                img = cv2.line(img, 
                    (int(speed_bar_xc), int(lbl_speed_tri_yc)), 
                    (int(speed_bar_xc), int(out_speed_tri_yc)), 
                    color=lbl_out_distance_color, thickness=fg_line_thickness
                    )

                # PLOT: Label Speed
                lbl_speed_tri_yc = int(speed_bar_y2 - speed_bar_h * lbl_speed)
                lbl_speed_tri_yc = min([lbl_speed_tri_yc, speed_bar_y2])
                lbl_speed_tri_yc = max([lbl_speed_tri_yc, speed_bar_y1])

                lbl_speed_tri_y1 = int(lbl_speed_tri_yc - speed_bar_w / 4)
                lbl_speed_tri_y2 = int(lbl_speed_tri_yc + speed_bar_w / 4)
                
                pt1 = (speed_bar_xc, lbl_speed_tri_yc)
                pt2 = (speed_bar_x2, lbl_speed_tri_y1)
                pt3 = (speed_bar_x2, lbl_speed_tri_y2)

                img = cv2.drawContours(img, 
                    [np.array([pt1, pt2, pt3], dtype=int)], 
                    contourIdx=-1, color=lbl_color, thickness=-1
                    )

                # PLOT: Output Speed
                out_speed_tri_yc = speed_bar_y2 - int(speed_bar_h * out_speed)
                out_speed_tri_yc = min([out_speed_tri_yc, speed_bar_y2])
                out_speed_tri_yc = max([out_speed_tri_yc, speed_bar_y1])

                out_speed_tri_y1 = int(out_speed_tri_yc - speed_bar_w / 4)
                out_speed_tri_y2 = int(out_speed_tri_yc + speed_bar_w / 4)
                
                pt1 = (speed_bar_xc, out_speed_tri_yc)
                pt2 = (speed_bar_x1, out_speed_tri_y1)
                pt3 = (speed_bar_x1, out_speed_tri_y2)

                img = cv2.drawContours(img, 
                    [np.array([pt1, pt2, pt3], dtype=int)], 
                    contourIdx=-1, color=out_color, thickness=-1
                    )
                



                

                # PLOT: Loss box
                try: speed_bar_x2
                except: speed_bar_x2 = 0
                
                loss_box_x1 = int(speed_bar_x2 + img_height * 0.05)
                loss_box_x2 = int(img_width - img_height * 0.05)
                loss_box_y1 = int(img_height * 0.75)
                loss_box_y2 = int(img_height * 0.95)

                loss_box_w = loss_box_x2 - loss_box_x1
                loss_box_h = loss_box_y2 - loss_box_y1

                loss_box_xc = loss_box_x1 + int(loss_box_w/2)
                loss_box_yc = loss_box_y1 + int(loss_box_h/2)

                img = cv2.rectangle(img, 
                    (loss_box_x1, loss_box_y1), 
                    (loss_box_x2, loss_box_y2), 
                    color=bg_color, thickness=bg_line_thickness
                    )

                # PLOT: Loss
                loss_dx = loss_box_w / batch_n
                
                pts = np.array([
                    [item for sublist in [[loss_box_x1 + x*loss_dx, loss_box_x1 + (x+1)*loss_dx] for x in range(batch_i + 1)] for item in sublist],
                    [loss_box_y2 - loss_box_h * x/loss_over_batches[-1] for x in loss_over_batches[:batch_i + 1] for _ in range(2)]
                    ], np.int32)
                pts = np.transpose(pts, (1, 0))
                pts = pts.reshape((-1, 1, 2))

                img = cv2.polylines(img, 
                    [pts], 
                    isClosed=False, color=loss_color, thickness=fg_line_thickness
                    )


                # PLOT: LEGEND
                img = cv2.putText(img, 
                    text=f'LBL',
                    org=(int(img_width * 0.8), int(img_height * 0.1)),
                    fontFace=cv2.FONT_HERSHEY_DUPLEX, 
                    fontScale=0.4 * (1 + img_width/720 - 1/4), 
                    color=lbl_color, thickness=text_line_thickness
                    )

                img = cv2.putText(img, 
                    text=f'OUT',
                    org=(int(img_width * 0.8), int(img_height * 0.2)),
                    fontFace=cv2.FONT_HERSHEY_DUPLEX, 
                    fontScale=0.4 * (1 + img_width/720 - 1/4), 
                    color=out_color, thickness=text_line_thickness
                    )

                img = cv2.putText(img, 
                    text=f'LOSS',
                    org=(int(loss_box_x1 + int(img_height * 0.02)), int(loss_box_y1 + (loss_box_y2 - loss_box_y1) * 0.5)),
                    fontFace=cv2.FONT_HERSHEY_DUPLEX, 
                    fontScale=0.4 * (1 + img_width/720 - 1/4), 
                    color=loss_color, thickness=text_line_thickness
                    )
                img = cv2.putText(img, 
                    text=f'{loss_over_batches[-1]:.6f}',
                    org=(int(loss_box_x2 - int(img_width * 0.4)), int(loss_box_y1 + (loss_box_y2 - loss_box_y1) * 0.9)),
                    fontFace=cv2.FONT_HERSHEY_DUPLEX, 
                    fontScale=0.4 * (1 + img_width/720 - 1/4), 
                    color=loss_color, thickness=text_line_thickness
                )

            
                video_writer.write(img.astype(np.uint8))



            














"""def plot_loss_over_epochs(train_loss_over_epochs, validation_loss_over_epochs):
    plt.figure(figsize=(10, 7))
    plt.plot(train_loss_over_epochs, color='orange', label='Training loss')
    plt.plot(validation_loss_over_epochs, color='red', label='Validation loss')
    plt.xlabel('Epochs')
    plt.ylabel(config.LOSS_TYPE)
    plt.legend()
    plt.tight_layout()
    plt.savefig(f"{config.EXP_DIR}/loss_over_epochs.pdf")
    plt.close()

def plot_learn_rate_over_epochs(learning_rate_over_epochs):
    plt.figure(figsize=(10, 7))
    if config.LEARN_RATE_SCHEDULER_TYPE == 'ExponentialLR':
        plt.plot(learning_rate_over_epochs, color='yellow', 
            label=f'{config.LEARN_RATE_SCHEDULER_TYPE}, \gamma={config.LEARN_RATE_GAMMA}')
    else:
        raise ValueError(f"No implementation provided for 'LEARNING_RATE_SCHEDULER_TYPE={config.LEARN_RATE_SCHEDULER_TYPE}'.")
    plt.xlabel('Epochs')
    plt.ylabel('Learning Rate')
    plt.legend()
    plt.tight_layout()
    plt.savefig(f"{config.EXP_DIR}/learning_rate_over_epochs.pdf")
    plt.close()"""

def update_recordings_plot (recordings: Dict['str', List[float]]) -> None:
    plt.rcParams['text.usetex'] = True
    fig, ax1 = plt.subplots(figsize=(10, 7))
    color = 'tab:orange'
    ax1.set_xlabel('Epochs')
    ax1.set_ylabel(recordings['loss']['id'], color=color)
    plt.plot(recordings['loss']['train'], color=color, linestyle=':', label='Training')
    plt.plot(recordings['loss']['valid'], color=color, linestyle='-', label='Validation')
    ax1.tick_params(axis='y', labelcolor=color)
    #ax1.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.3e'))
    
    ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis
    color = 'tab:blue'
    ax2.set_ylabel('Learning Rate', color=color)  # we already handled the x-label with ax1
    ax2.plot(recordings['learn_rate']['values'], color=color, 
        label=f"{recordings['learn_rate']['id']}, $\gamma={recordings['learn_rate']['gamma']}$")
    ax2.tick_params(axis='y', labelcolor=color)
    ax2.yaxis.set_major_formatter(mtick.FormatStrFormatter('%.1e'))
    
    ax1.legend(loc='lower left')
    ax2.legend(loc='upper right')
    fig.tight_layout()  # otherwise the right y-label is slightly clipped    
    plt.savefig(config.get_experiment_dpath()/'recordings_lin_yscale.pdf')

    ax1.set_yscale('log')
    ax2.set_yscale('log')
    plt.savefig(config.get_experiment_dpath()/'recordings_log_yscale.pdf')
    plt.close()




"""def get_summary_keys():
    S1C1 = [
        'ANN', 
        'DATA', 
        'LEARN',
    ]
    S1C2 = [
        ['CNN', 'RNN', 'FC'],
        ['RUNS', 'INPUTS', 'LABELS'],
        ['Batch Size', '#Epochs', 'Optimizer', 'L RATE', 'LOSS'],
        ['EPOCHS', 'TRAIN_LOSS', 'VALID_LOSS', 'LEARN_RATE']
    ]
    S1C3 = [
        [
            ['Model', 'Pretrained', 'Frozen', 'Dropout'], 
            ['Type', 'Hidden Size', '#Layers', 'Dropout'], 
            ['Output Size', '#Layers', 'Act Fct', 'Dropout']
        ],
        [
            #['Mode', 'FIG8', 'ITL', '#Train/Sc', '#Valid/Sc', '#Images'], 
            ['Mode', 'FIG8', 'ITL', '#Images'], 
            ['Img Resize', 'IMU', 'Seq Length', 'Seq Step'], 
            ['Waypoint', 'Speed', 'Ctrl Cmd']
        ],
        [
            [''], 
            [''], 
            [''],
            ['Init', 'Scheduler', 'Gamma'], 
            ['Type', 'Sp/Wp Ratio']
        ],
    ]
    S2R1 = ['Epochs', 'T Loss', 'V Loss', 'L Rate']
    
    return S1C1, S1C2, S1C3, S2R1"""



"""def save_summary(summary):
    summary_df = pd.DataFrame.from_dict(summary)
    summary_df.to_csv(config.EXP_DIR/'summary.csv')

    with open(config.EXP_DIR/'summary.json', 'w') as file:
        file.write(json.dumps(summary)) # use `json.loads` to do the reverse

    wb = openpyxl.Workbook()
    wb_sheet_1 = wb.active
    wb_sheet_1.title = 'Parameters'
    wb_sheet_2 = wb.create_sheet(title="Results")
    
    S1C1, S1C2, S1C3, S2R1 = get_summary_keys()

    row_counter = 0
    
    for i1 in range(len(S1C1)):

        row_counter += 1

        wb_sheet_1.cell(
            column=1, 
            row=row_counter, 
            value=S1C1[i1]
        )
        
        for i2 in range(len(S1C2[i1])):

            wb_sheet_1.cell(
                column=2, 
                row=row_counter, 
                value=S1C2[i1][i2]
            )
            
            for i3 in range(len(S1C3[i1][i2])):
            
                wb_sheet_1.cell(
                    column=3, 
                    row=row_counter, 
                    value=S1C3[i1][i2][i3]
                )

                wb_sheet_1.cell(
                    column=4, 
                    row=row_counter, 
                    value=summary[ S1C1[i1] ][ S1C2[i1][i2] ][ S1C3[i1][i2][i3] ]
                )

                row_counter +=1


    for i in range(len(S2R1)):
        wb_sheet_2.cell(
            column=1 + i, 
            row=1, 
            value=S2R1[i]
        )

    for i in range(len(S2R1)):
        for j in range(len(summary['RESULTS'][ S2R1[i] ])):
            wb_sheet_2.cell(
                column=i + 1, 
                row=j + 2, 
                value=summary['RESULTS'][ S2R1[i] ][ j ]
            )
    
    wb_file = str(config.EXP_DIR/'summary.xlsx')
    wb.save(filename=wb_file)"""




def save_checkpoint (
    epoch_i: int, 
    model: torch.nn.Module, 
    optimizer: torch.optim.Optimizer, 
    loss: torch.nn.Module,
    fpath: pathlib.Path
) -> None:
    torch.save(
        {
            'epoch': epoch_i,
            'model_state_dict': model.state_dict(),
            'optimizer_state_dict': optimizer.state_dict(),
            'loss': loss,
        }, 
        fpath
    )

def export_to_json (
    configuration: Dict, 
    fpath: pathlib.Path
) -> None:
    with open(fpath, 'w') as file:
        file.write(json.dumps(configuration, sort_keys=True, indent=4))


"""def save_model_as_script_module (
    model: torch.nn.Module, 
    script_mode: str,
    fpath: pathlib.Path,
) -> None:

    # 0] For inference, batch size and sequence length are 1
    bs = 1; sl = 1

    # 1] Converting PyTorch Model to Torch Script
    if script_mode == 'tracing':
        img = torch.rand(bs, sl, 3, config.IMAGE_HEIGHT_RESIZED, config.IMAGE_WIDTH_RESIZED)
        imu = torch.rand(bs, sl, len(config.COLS_INPUT_IMU))
        h = torch.rand(config.GRU_NUM_LAYERS, bs, config.GRU_HIDDEN_SIZE)
        example_input = [img, imu, h]
        script_module = torch.jit.trace(model, *example_input)
    elif script_mode == 'annotation':
        script_module = torch.jit.script(model)
    else:
        raise ValueError(f"Parameter `script_mode` must be either 'tracing' or 'annotation'.")

    # 2] Serializing Script Module to a File 
    script_module.save(str(config.EXP_DIR/"model_script_module.pt"))
"""


"""def update_experiment_overview():

    S1C1, S1C2, S1C3, S2R1 = get_summary_keys()

    exp_dirs = sorted(
        [
            x for x in config.OUTPUT_DIR.iterdir()\
                if x.is_dir() and x.name[10:13] == "___"
        ]
    )

    exp_n = len(exp_dirs)
    
    exp_dict = dict()
    wb = openpyxl.Workbook()
    ws_1 = wb.active
    ws_1.title = 'Parameters'
    ws_2 = wb.create_sheet(title="Results")


    
    
    
    for dir_i, dir in enumerate(exp_dirs):
        
        try:
            summary_json_file = dir/'summary.json'
            with open(summary_json_file, 'r') as j:
                exp_dict[dir.name] = json.loads(j.read())
        except:
            continue

        col_i1 = dir_i + 4
        row_i1 = 1
        row_i2 = 1
        
        # Experiment IDs
        ws_1.cell(
            column=col_i1, 
            row=row_i1, 
            value=dir.name
        )
        row_i1 += 1

        ws_2.cell(
            column=3 * dir_i + 3,
            row=row_i2, 
            value=dir.name
        )
        row_i2 += 1




        for i1 in range(len(S1C1)):
            if dir_i == 0: ws_1.cell(
                column=1, 
                row=row_i1, 
                value=S1C1[i1]
            )
            for i2 in range(len(S1C2[i1])):
                if dir_i == 0: ws_1.cell(
                    column=2, 
                    row=row_i1,
                    value=S1C2[i1][i2]
                )
                for i3 in range(len(S1C3[i1][i2])):
                    if dir_i == 0: ws_1.cell(
                        column=3, 
                        row=row_i1, 
                        value=S1C3[i1][i2][i3]
                    )
                    try:
                        ws_1.cell(
                            column=col_i1, 
                            row=row_i1, 
                            value=exp_dict[dir.name][ S1C1[i1] ][ S1C2[i1][i2] ][ S1C3[i1][i2][i3] ]
                        )
                    except KeyError:
                        ws_1.cell(
                            column=col_i1, 
                            row=row_i1, 
                            value='_'
                        )
                    row_i1 += 1
        for i in range(1, len(S2R1)):
            ws_2.cell(
                column=3 * dir_i + 2 + i,
                row=2,
                value=S2R1[i]
            )
            for j in range(len(exp_dict[dir.name]['RESULTS'][ S2R1[i] ])):
                ws_2.cell(
                    column=3 * dir_i + 2 + i, 
                    row=j + 3, 
                    value=exp_dict[dir.name]['RESULTS'][ S2R1[i] ][ j ]
                )

        
    wb_file = config.OUTPUT_DIR/'experiments.xlsx'
    wb_file.unlink(missing_ok=True)

    wb.save(filename=str(wb_file))

    exp_df = pd.DataFrame.from_dict(exp_dict)
    exp_df.to_csv(config.OUTPUT_DIR/'experiments.csv')

    with open(config.OUTPUT_DIR/'experiments.json', 'w') as file:
        file.write(json.dumps(exp_dict))
"""
    




if __name__ == "__main__":
    pass
    #update_experiment_overview()